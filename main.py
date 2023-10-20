import sys
import os
import datetime
import random
import ctypes

import openpyxl
import py_win_keyboard_layout
import csv
import Algoritm
import time
from openpyxl import load_workbook

from PyQt5.QtCore import QRegExp, Qt
from PyQt5.QtGui import QRegExpValidator, QFont, QPixmap
from PyQt5 import QtWidgets, QtGui
from PyQt5.QtWidgets import QApplication, QMainWindow, QLineEdit, QMessageBox, QLabel, QFileDialog, QSplashScreen

# from IPython.external.qt_for_kernel import QtGui


def resource_path(relative):
    if hasattr(sys, '_MEIPASS'):
        return os.path.join(sys._MEIPASS, relative)
    else:
        return os.path.join(os.path.abspath("."), relative)


class Instruction(QMainWindow):
    def edit_back_image(self, count):
        self.setStyleSheet("#MainWindow{border-image:url(Wpora_phon.png)}")

    def __init__(self):
        super().__init__()

class SplashScreen(QSplashScreen):
    def __init__(self):
        super(QSplashScreen, self).__init__()
        self.setWindowFlag(Qt.FramelessWindowHint)
        pixmap = QPixmap("vs.png")
        self.setPixmap(pixmap)
    def progress(self):
        for i in range(3):
            time.sleep(1)
class Window(QMainWindow):
    def russkladka_rus(self, tmp):  # раскладка
        u = ctypes.windll.LoadLibrary("user32.dll")
        pf = getattr(u, "GetKeyboardLayout")
        if tmp == 1:  # на русский
            if hex(pf(0)) == '0x4090409':
                py_win_keyboard_layout.change_foreground_window_keyboard_layout(0x4190419)
        else:  # на engl
            if hex(pf(0)) == '0x4190419':
                py_win_keyboard_layout.change_foreground_window_keyboard_layout(0x4090409)

    def button_random_click(self):
        if self.first_random or self.end_random_pass:
            self.first_random = False
            value = random.randint(2, 6)
            if value == 2:
                self.table_1_1.setText(
                    str(random.choice(Algoritm.English_arr[1:-1])) + '-' + str(random.randint(1, 60)) + '-' + str(
                        random.choice(random.choice(Algoritm.table_rimnumber))))
            elif value == 3:
                self.table_1_1.setText(
                    str(random.choice(Algoritm.English_arr[1:-1])) + '-' + str(random.randint(1, 60)) + '-' + str(
                        random.choice(Algoritm.Russian_Arr)))
            elif value == 4:
                self.table_1_1.setText(
                    str(random.choice(Algoritm.English_arr[1:-1])) + '-' + str(random.randint(1, 60)) + '-' + str(
                        random.randint(1, 144)))
            elif value == 5:
                self.table_1_1.setText(
                    str(random.choice(Algoritm.English_arr[1:-1])) + '-' + str(random.randint(1, 60)) + '-' + str(
                        random.randint(1, 144)) + '-' + str(random.choice(Algoritm.Russian_Arr)))
            elif value == 6:
                self.table_1_1.setText(
                    str(random.choice(Algoritm.English_arr[1:-1])) + '-' + str(random.randint(1, 60)) + '-' + str(
                        random.randint(1, 144)) + '-' + str(random.choice(Algoritm.Russian_Arr)) + '-' + str(
                        random.choice(Algoritm.Russian_Arr)).lower())
            self.require_random = False
            self.hint_text.adjustSize()
        else:
            self.restart_text.show()
            self.restart_value.show()
            self.require_random = True
            self.button_OK.show()
            self.russkladka_rus(2)  # раскладка на англ
        self.table_1_1.setReadOnly(True)

    def button_instruction_click(self):
        self.instruction.edit_back_image(self.count_isntruction)
        self.instruction.showMaximized()

    def button_test_finish_click(self):
        self.count_isntruction = 4
        self.button_test_finish.hide()
        self.button_new_student.show()
        self.button_of_restart_variant.show()
        self.table_0_0.setReadOnly(True)
        self.table_0_1.setReadOnly(True)
        self.table_0_2.setReadOnly(True)
        self.table_1_0.setReadOnly(True)
        self.table_1_2.setReadOnly(True)
        self.table_2_0.setReadOnly(True)
        self.table_2_1.setReadOnly(True)
        self.table_2_2.setReadOnly(True)
        self.table_2_3.setReadOnly(True)
        count = 10
        style_field_red = "border: 2px solid grey; background: red;  color: black; border-radius: 15px; "
        style_field_green = "border: 2px solid grey; background: green;  color: black; border-radius: 15px; "

        if self.table_0_0.text() == self.Table[0][0]:
            self.table_0_0.setStyleSheet(style_field_green)
        else:
            self.table_0_0.setStyleSheet(style_field_red)
            count -= 1

        if self.table_0_1.text() == self.Table[0][1]:
            self.table_0_1.setStyleSheet(style_field_green)
        else:
            self.table_0_1.setStyleSheet(style_field_red)
            count -= 1

        if self.table_0_2.text() == self.Table[0][2]:
            self.table_0_2.setStyleSheet(style_field_green)
        else:
            count -= 1
            self.table_0_2.setStyleSheet(style_field_red)

        if self.table_1_0.text() == self.Table[1][0]:
            self.table_1_0.setStyleSheet(style_field_green)
        else:
            count -= 1
            self.table_1_0.setStyleSheet(style_field_red)
        if self.table_1_2.text() == self.Table[1][2]:
            self.table_1_2.setStyleSheet(style_field_green)
        else:
            count -= 1
            self.table_1_2.setStyleSheet(style_field_red)

        if self.table_2_0.text() == self.Table[2][0]:
            self.table_2_0.setStyleSheet(style_field_green)
        else:
            self.table_2_0.setStyleSheet(style_field_red)
            count -= 1

        if self.table_2_1.text() == self.Table[2][1]:
            self.table_2_1.setStyleSheet(style_field_green)
        else:
            count -= 1
            self.table_2_1.setStyleSheet(style_field_red)

        if self.table_2_2.text() == self.Table[2][2]:
            self.table_2_2.setStyleSheet(style_field_green)
        else:
            count -= 1
            self.table_2_2.setStyleSheet(style_field_red)

        temp_str = self.table_2_3.text().replace(' ', '')

        if temp_str == self.Table[2][3]:
            self.table_2_3.setStyleSheet(style_field_green)
        else:
            count -= 1
            self.table_2_3.setStyleSheet(style_field_red)

        othenka = '5'
        if count == 9:
            othenka = '4'
        elif count == 8:
            othenka = '3'
        elif count < 8:
            othenka = '2'
        QMessageBox.about(self, "Внимание", "Тест завершен, вы набрали " +
                          str(count) + "/"
                          + str(10) + " баллов\n Ваша оцека: " + othenka)
        self.button_restart.show()

        # print(self.dirlist)
        out_directory = os.path.dirname(__file__)
        now = datetime.datetime.now()
        tm = str(now.hour) + ":" + str(now.minute)
        dt = str(now.day) + "-" + str(now.month) + "-" + str(int(now.year) - 2000)
        if os.path.isfile("Номенклатура_результаты.xlsx"):
            wb = load_workbook("Номенклатура_результаты.xlsx")
            ws = wb['Sheet']
            ws.append([self.FIO_value.text(), self.vzvod_value.text(), count, othenka, tm, dt])
            wb.save("Номенклатура_результаты.xlsx")
            wb.close()
        else:
            wb = openpyxl.Workbook()
            ws = wb['Sheet']
            ws.append(["ФИО", "Взвод", "Количество баллов", "Оценка", "Время завершения", "Дата"])
            ws.append([self.FIO_value.text(), self.vzvod_value.text(), count, othenka, tm, dt])
            wb.save("Номенклатура_результаты.xlsx")
            wb.close()
        if os.path.isfile("Path.csv"):
            with open("Path.csv", mode="r") as file:
                reader = csv.reader(file)
                spisok = list(reader)
                stroka = str(" ".join(spisok[0])) + "/Номенклатура_результаты.xlsx"
                if os.path.isfile(stroka):
                    wb = load_workbook(stroka)
                    ws = wb['Sheet']
                    ws.append([self.FIO_value.text(), self.vzvod_value.text(), count, othenka, tm, dt])
                    wb.save(stroka)
                    wb.close()
                else:
                    if os.path.isdir(str("".join(spisok[0]))):
                            wb = openpyxl.Workbook()
                            ws = wb['Sheet']
                            ws.append(["ФИО", "Взвод", "Количество баллов", "Оценка", "Время завершения", "Дата"])
                            ws.append([self.FIO_value.text(), self.vzvod_value.text(), count, othenka, tm, dt])
                            wb.save(stroka)
                            wb.close()
            file.close()
        else:
            self.dirlist = QFileDialog.getExistingDirectory(self, "Выбрать папку", ".")
            print(self.dirlist)
            with open("Path.csv", mode="w+") as f:
                writer = csv.writer(f, dialect='excel')
                writer.writerow([self.dirlist])
                f.close()
            os.chdir(self.dirlist)
            if os.path.isfile("Номенклатура_результаты.xlsx"):
                wb = load_workbook("Номенклатура_результаты.xlsx")
                ws = wb['Sheet']
                ws.append([self.FIO_value.text(), self.vzvod_value.text(), count, othenka, tm, dt])
                wb.save("Номенклатура_результаты.xlsx")
                wb.close()
            else:
                wb = openpyxl.Workbook()
                ws = wb['Sheet']
                ws.append(["ФИО", "Взвод", "Количество баллов", "Оценка", "Время завершения", "Дата"])
                ws.append([self.FIO_value.text(), self.vzvod_value.text(), count, othenka, tm, dt])
                wb.save("Номенклатура_результаты.xlsx")
                wb.close()

    def button_OK_click(self):
        if self.restart_value.text() != 'yjvtyrkfnehf':
            QMessageBox.about(self, "Ошибка", "Неверный пароль")
        else:
            self.label_wpora2.hide()
            self.button_instruction.hide()
            self.russkladka_rus(2)
            self.label_wpora.hide()
            if self.require_random:
                self.end_random_pass = True
                self.require_random = False
                self.first_random = True
                self.restart_value.clear()
                self.restart_value.hide()
                self.restart_text.hide()
                self.button_OK.hide()
                self.button_random_click()

                self.button_of_restart_variant.hide()

            else:
                self.end_random_pass = True
                self.button_OK.hide()
                self.restart_value.hide()
                self.restart_text.hide()
                self.restart_value.clear()
                self.table_0_0.clear()
                self.table_0_1.clear()
                self.table_0_2.clear()
                self.table_1_0.clear()
                if self.save_1_1 == False:
                    self.table_1_1.clear()
                self.table_1_2.clear()
                self.table_2_0.clear()
                self.table_2_1.clear()
                self.table_2_2.clear()
                self.table_2_3.clear()
                self.mashtab_text.hide()
                self.symbol1k.hide()
                self.table_0_0.hide()
                self.table_0_1.hide()
                self.table_0_2.hide()
                self.table_1_0.hide()
                self.table_1_1.hide()
                self.table_1_2.hide()
                self.table_2_0.hide()
                self.table_2_1.hide()
                self.table_2_2.hide()
                self.table_2_3.hide()
                self.table_0_0.setReadOnly(False)
                self.table_0_1.setReadOnly(False)
                self.table_0_2.setReadOnly(False)
                self.table_1_0.setReadOnly(False)
                self.table_1_1.setReadOnly(False)
                self.table_1_2.setReadOnly(False)
                self.table_2_0.setReadOnly(False)
                self.table_2_1.setReadOnly(False)
                self.table_2_2.setReadOnly(False)
                self.table_2_3.setReadOnly(False)
                self.button_of_restart_variant.hide()
                # self.button_start.show()
                self.first_random = True
                self.Table[0][0] = '00'
                self.enter_button()

    def button_restart_click(self):
        self.button_new_student.hide()
        self.button_restart.hide()
        self.button_of_restart_variant.hide()
        self.restart_text.show()
        self.restart_value.show()
        self.button_OK.show()
        self.russkladka_rus(2)  # раскладка на англ
        ##self.label_wpora.hide()


    def button_start_click(self):
        self.button_instruction.show()
        self.label_wpora2.show()
        self.save_1_1 = False
        self.button_OK.hide()
        self.restart_value.clear()
        self.restart_value.hide()
        self.restart_text.hide()
        self.require_random = False
        self.Table = Algoritm.algoritm(str(self.table_1_1.text()))
        if self.Table[0][0] == '00':
            QMessageBox.about(self, "Внимание", "Элемент задан неправильно")
        else:
            self.count_isntruction = 3
            self.hint_text.setText("\n  Пример ввода\n  A(ю.п.)-40-B")
            self.hint_text.adjustSize()
            self.hint_text.show()
            self.symbol1k.setText("1:")
            self.mashtab_text.setText("Укажите масштаб!")
            self.mashtab_text.setAlignment(Qt.AlignCenter)
            self.label_wpora.show()
            self.mashtab_text.adjustSize()

            self.symbol1k.show()
            self.mashtab_text.show()
            style_field = "border: 2px solid grey; background-color:white; color:black;  border-radius: 15px;"
            style_field_button = "border: 2px solid grey; background-color:white; color:black;  border-radius: 10px;"
            self.table_0_0.setStyleSheet(style_field)
            self.table_0_1.setStyleSheet(style_field)
            self.table_0_2.setStyleSheet(style_field)
            self.table_1_0.setStyleSheet(style_field)
            self.table_1_1.setStyleSheet(style_field)
            self.table_1_2.setStyleSheet(style_field)
            self.table_2_0.setStyleSheet(style_field)
            self.table_2_1.setStyleSheet(style_field)
            self.table_2_2.setStyleSheet(style_field)
            self.table_2_3.setStyleSheet(style_field)
            self.table_0_0.show()
            self.table_0_1.show()
            self.table_0_2.show()
            self.table_1_0.show()
            self.table_1_1.setReadOnly(True)
            self.table_1_2.show()
            self.table_2_0.show()
            self.table_2_1.show()
            self.table_2_2.show()
            self.table_2_3.show()
            self.label_wpora.show()
            print(self.Table)
            # self.table_0_0.setText(Table[0][0])
            self.button_start.hide()
            self.button_test_finish.show()
            self.russkladka_rus(2)

    def enter_button_continue(self):
        self.count_isntruction = 2
        self.table_1_1.show()
        self.button_start.show()

        self.button_normal.hide()
        self.hint_text.setText("\nВведите число номер поля\n Например: A(ю.п.)-40-B")
        self.hint_text.adjustSize()


    def enter_button(self):
        if self.vzvod_value.text() == '' or self.FIO_value.text() == '':
            QMessageBox.about(self, "Предупреждение", "Заполните все поля")
        else:
            self.count_isntruction = 2
            self.FIO_text.setHidden(True)  # hide a vhod elemnts
            self.vzvod_text.setHidden(True)
            self.FIO_value.setHidden(True)
            self.vzvod_value.setHidden(True)
            self.button_enter.hide()
            self.table_1_1.show()
            # self.hint_text.setText("\nВведите число номер поля\n Например: A(ю.п.)-40-B")
            self.hint_text.setText("\nЗадайте Вариант!")
            self.Identification_text.setText(
                str(self.vzvod_value.text()) + " " + str(self.FIO_value.text()))  # identification text set
            self.hint_text.adjustSize()
            self.Identification_text.adjustSize()  # update size for text
            # self.table_1_1.show()
            # self.button_start.show()
            self.button_normal.show()

            self.button_restart.hide()
            self.button_start.show()
            self.russkladka_rus(2)

    def button_of_restart_variant_click(self):
        self.count_isntruction = 2
        tmp = self.table_1_1.text()
        self.button_restart.hide()
        self.button_of_restart_variant.hide()
        self.button_new_student.hide()
        self.save_1_1 = True
        self.restart_value_variant = True
        self.button_OK_2.show()
        self.restart_text.show()
        self.restart_value.show()

        self.russkladka_rus(2)
        self.button_instruction.show()
        ######################################################################################

    def button_OK_2_click(self):
        if self.restart_value.text() != 'yjvtyrkfnehf':
            QMessageBox.about(self, "Ошибка", "Неверный пароль")
        else:
            self.label_wpora2.hide()
            self.button_instruction.hide()
            self.russkladka_rus(2)
            if self.restart_value_variant:
                self.table_0_0.setReadOnly(False)
                self.table_0_1.setReadOnly(False)
                self.table_0_2.setReadOnly(False)
                self.table_1_0.setReadOnly(False)
                self.table_1_2.setReadOnly(False)
                self.table_2_2.setReadOnly(False)
                self.table_2_1.setReadOnly(False)
                self.table_2_0.setReadOnly(False)
                self.table_2_3.setReadOnly(False)

                self.table_0_0.clear()
                self.table_0_1.clear()
                self.table_0_2.clear()
                self.table_1_0.clear()
                self.table_1_2.clear()
                self.table_2_2.clear()
                self.table_2_1.clear()
                self.table_2_0.clear()
                self.table_2_3.clear()
                self.restart_value_variant = False
                self.button_start_click()
                self.button_OK_2.hide()
    def button_new_student_click(self):
        self.label_wpora2.hide()
        self.button_instruction.hide()
        self.count_isntruction = 1
        self.button_new_student.hide()
        self.button_restart.hide()
        self.button_of_restart_variant.hide()
        self.label_wpora.hide()
        self.save_1_1 = False
        self.restart_value.hide()
        self.restart_text.hide()
        self.restart_value.clear()
        self.table_0_0.clear()
        self.table_0_1.clear()
        self.table_0_2.clear()
        self.table_1_0.clear()
        self.table_1_1.clear()
        self.table_1_2.clear()
        self.table_2_0.clear()
        self.table_2_1.clear()
        self.table_2_2.clear()
        self.table_2_3.clear()
        self.Table[0][0] = "00"  ######################################
        self.table_0_0.setReadOnly(False)
        self.table_0_1.setReadOnly(False)
        self.table_0_2.setReadOnly(False)
        self.table_1_0.setReadOnly(False)
        self.table_1_2.setReadOnly(False)
        self.table_2_0.setReadOnly(False)
        self.table_2_1.setReadOnly(False)
        self.table_2_2.setReadOnly(False)
        self.table_2_3.setReadOnly(False)
        self.mashtab_text.hide()
        self.symbol1k.hide()
        self.table_0_0.hide()
        self.table_0_1.hide()
        self.table_0_2.hide()
        self.table_1_0.hide()
        self.table_1_1.hide()
        self.table_1_2.hide()
        self.table_2_0.hide()
        self.table_2_1.hide()
        self.table_2_2.hide()
        self.table_2_3.hide()
        self.hint_text.hide()
        self.button_normal.hide()
        self.FIO_text.show()
        self.vzvod_text.show()
        self.button_enter.show()
        self.FIO_value.show()
        self.vzvod_value.show()
        self.FIO_value.clear()
        self.vzvod_value.clear()
        self.Identification_text.clear()
        self.table_1_1.setReadOnly(False)
        self.russkladka_rus(1)
        self.end_random_pass = False
        self.first_random = True  # для певого нажатия
        self.request_random = False  # для запроса рандома
        self.russkladka_rus(1)

    def __init__(self, instruction) -> None:
        """"create a window object"""
        self.instruction = instruction
        super(Window, self).__init__()
        self.count_isntruction = 1  # для инструкции счетчик
        self.dirlist = None
        self.save_1_1 = False
        self.restart_value_variant = False
        self.label_wpora = QLabel(self)
        self.pixmap_wpora = QPixmap(resource_path('Wpora.png'))
        self.label_wpora.setPixmap(self.pixmap_wpora)
        self.label_wpora.resize(self.pixmap_wpora.width(), self.pixmap_wpora.height())
        self.label_wpora.hide()
        self.end_random_pass = False
        self.first_random = True  # для певого нажатия
        self.request_random = False  # для запроса рандома
        self.label = QLabel(self)  ##########
        self.pixmap = QPixmap(resource_path('Gerb___.png'))  ##########
        self.pixmap = self.pixmap.scaled(500, 500)
        self.label.setPixmap(self.pixmap)  #############
        self.label.resize(self.pixmap.width(), self.pixmap.height())  ###############
        self.label.move(-40, -10)  ############
        self.label.show()  #############
        self.clicked_finish = bool
        self.symbol1k = QtWidgets.QLabel(self)
        self.button_start = QtWidgets.QPushButton(self)
        self.button_test_finish = QtWidgets.QPushButton(self)  # button finish test
        self.button_enter = QtWidgets.QPushButton(self)  # button enter in system
        self.button_restart = QtWidgets.QPushButton(self)
        self.button_OK = QtWidgets.QPushButton(self)
        self.button_OK_2 = QtWidgets.QPushButton(self)
        self.button_of_restart_variant = QtWidgets.QPushButton(self)
        self.button_new_student = QtWidgets.QPushButton(self)
        self.button_normal = QtWidgets.QPushButton(self)
        self.button_instruction = QtWidgets.QPushButton(self)


        self.label_wpora2 = QLabel(self)
        self.pixmap_wpora2 = QPixmap(resource_path('Wpora2.png'))
        self.label_wpora2.setPixmap(self.pixmap_wpora2)
        self.label_wpora2.resize(self.pixmap_wpora2.width(), self.pixmap_wpora2.height())
        self.label_wpora2.hide()

        self.button_enter.setFocusPolicy(Qt.NoFocus)
        self.button_start.setFocusPolicy(Qt.NoFocus)
        self.button_normal.setFocusPolicy(Qt.NoFocus)
        self.button_instruction.setFocusPolicy(Qt.NoFocus)

        self.mashtab_text = QtWidgets.QLabel(self)
        self.hint_text = QtWidgets.QLabel(self)  # label text question
        self.hint_text.setFont(QFont('Times', 12))
        self.vzvod_text = QtWidgets.QLabel(self)  # label text enter a vzvod
        self.vzvod_text.setFont(QFont('Times', 12))
        self.FIO_text = QtWidgets.QLabel(self)  # label yext enter a fio
        self.FIO_text.setFont(QFont('Times', 12))
        self.restart_text = QtWidgets.QLabel(self)
        self.restart_text.setFont(QFont('Times', 8))
        self.Identification_text = QtWidgets.QLabel(self)  # label identificator
        self.Identification_text.setFont(QFont('Times', 16))
        self.table_0_0 = QtWidgets.QLineEdit(self)
        self.table_0_1 = QtWidgets.QLineEdit(self)
        self.table_0_2 = QtWidgets.QLineEdit(self)
        self.table_1_0 = QtWidgets.QLineEdit(self)
        self.table_1_1 = QtWidgets.QLineEdit(self)
        self.table_1_2 = QtWidgets.QLineEdit(self)
        self.table_2_0 = QtWidgets.QLineEdit(self)
        self.table_2_1 = QtWidgets.QLineEdit(self)
        self.table_2_2 = QtWidgets.QLineEdit(self)
        self.table_2_3 = QtWidgets.QLineEdit(self)
        self.restart_value = QtWidgets.QLineEdit(self)
        self.restart_value.setEchoMode(QtWidgets.QLineEdit.Password)  ###########################
        self.FIO_value = QLineEdit(self)  # field fio value
        self.FIO_value.setFont(QFont('Times', 12))
        self.vzvod_value = QLineEdit(self)  # field vzvod value
        self.vzvod_value.setFont(QFont('Times', 12))
        reg_table = QRegExp("[A-Z\(ю\.п\.)]{,7}-[0-9]{,2}-[А-я,0-9,A-Z]{,5}-[А-Г]{,1}-[а-г]{,1}")
        input_validator_table = QRegExpValidator(reg_table, self.table_1_1)
        self.table_1_1.setValidator(input_validator_table)
        reg_vzvod = QRegExp("[0-9,-]{,15}")
        reg_FIO = QRegExp("[А-Яа-я,.- ]{,60}")
        input_validator = QRegExpValidator(reg_vzvod, self.vzvod_value)
        input_validator_FIO = QRegExpValidator(reg_FIO, self.FIO_value)
        self.vzvod_value.setValidator(input_validator)
        self.vzvod_value.setPlaceholderText(" Только цифры")
        self.FIO_value.setPlaceholderText(" Только русские буквы")
        self.FIO_value.setValidator(input_validator_FIO)
        self.button_start.hide()
        self.button_test_finish.hide()
        self.button_OK.hide()
        self.button_OK_2.hide()
        self.button_new_student.hide()
        # ________edit

        self.FIO_text.setText(" Введите ФИО: ")
        self.vzvod_text.setText(" Номер взвода: ")
        self.button_enter.setText(" Войти в систему ")
        self.button_start.setText(" Начать тест ")
        self.button_test_finish.setText(" Завершить тест ")
        self.button_restart.setText('Повторить')
        self.restart_text.setText(" Введите пароль ")
        self.button_of_restart_variant.setText(" Начать вариант заново ")
        self.button_OK.setText("ОК")
        self.button_OK_2.setText("ОК")
        self.button_new_student.setText(" Смена пользователя ")
        self.button_instruction.setText(" ")
        self.button_normal.setText(" Получить задание ")

        self.restart_value.move(670, 450)
        self.restart_text.move(670, 400)
        self.restart_text.setAlignment(Qt.AlignCenter)
        self.restart_text.setStyleSheet('color: rgb(255, 255, 255);')
        self.FIO_text.move(220, 220)  # sdvig elements
        self.vzvod_text.move(220, 200)
        self.FIO_value.move(320, 200)
        self.button_enter.move(320, 280)
        self.vzvod_value.move(320, 240)
        self.hint_text.move(290, 50)
        self.hint_text.setAlignment(Qt.AlignCenter)
        self.button_test_finish.move(680, 550)
        self.Identification_text.move(10, 0)
        self.symbol1k.move(607, 217)
        self.symbol1k.setFont(QFont('Times', 12))
        self.mashtab_text.setFont(QFont('Times', 12))
        self.table_1_1.setPlaceholderText("A(ю.п.)-40-144-Г-г")
        self.table_0_0.move(120, 187)
        self.table_0_1.move(285, 187)
        self.table_0_2.move(446, 187)
        self.table_1_0.move(120, 250)
        self.table_1_1.move(285, 250)
        self.table_1_2.move(446, 250)
        self.table_2_0.move(120, 317)
        self.table_2_1.move(285, 317)
        self.table_2_2.move(446, 317)
        self.table_2_3.move(640, 210)
        self.button_start.move(680, 550)
        self.mashtab_text.move(615, 190)
        self.button_restart.move(670, 550)
        self.button_OK.move(670, 550)
        self.button_OK_2.move(670, 550)
        self.button_of_restart_variant.move(300, 410)

        self.button_normal.move(285, 187)

        self.button_OK.adjustSize()
        self.button_OK_2.adjustSize()
        self.restart_text.adjustSize()
        self.button_restart.adjustSize()
        self.FIO_text.adjustSize()  # update size element
        self.vzvod_text.adjustSize()
        self.button_enter.adjustSize()
        self.button_start.adjustSize()
        self.button_test_finish.adjustSize()
        self.button_of_restart_variant.adjustSize()
        self.button_new_student.setFixedSize(100, 20)
        self.restart_value.setFixedSize(150, 50)
        self.FIO_value.setFixedSize(250, 40)  # fixed size
        self.vzvod_value.setFixedSize(150, 40)
        self.table_0_0.setFixedSize(150, 50)
        self.table_0_1.setFixedSize(150, 50)
        self.table_0_2.setFixedSize(150, 50)
        self.table_1_0.setFixedSize(150, 50)
        self.table_1_1.setFixedSize(150, 50)
        self.table_1_2.setFixedSize(150, 50)
        self.table_2_0.setFixedSize(150, 50)
        self.table_2_1.setFixedSize(150, 50)
        self.table_2_2.setFixedSize(150, 50)
        self.table_2_3.setFixedSize(150, 50)
        self.button_of_restart_variant.setFixedSize(150, 30)
        self.button_normal.setFixedSize(150, 50)
        self.button_restart.setFixedSize(100, 30)
        self.button_new_student.setFixedSize(150, 30)
        self.button_OK.hide()
        self.restart_value.hide()
        self.restart_text.hide()
        self.table_0_0.hide()
        self.table_0_1.hide()
        self.table_0_2.hide()
        self.table_1_0.hide()
        self.table_1_1.hide()
        self.table_1_2.hide()
        self.table_2_0.hide()
        self.table_2_1.hide()
        self.table_2_2.hide()
        self.table_2_3.hide()
        self.button_of_restart_variant.hide()
        self.button_normal.hide()
        self.button_instruction.hide()

        self.button_restart.hide()
        self.button_instruction.clicked.connect(self.button_instruction_click)
        self.button_enter.clicked.connect(self.enter_button)  # relation button and function
        self.button_start.clicked.connect(self.button_start_click)
        self.button_test_finish.clicked.connect(self.button_test_finish_click)
        self.button_restart.clicked.connect(self.button_restart_click)
        self.button_OK.clicked.connect(self.button_OK_click)
        self.button_OK_2.clicked.connect(self.button_OK_2_click)
        self.button_of_restart_variant.clicked.connect(self.button_of_restart_variant_click)
        self.button_new_student.clicked.connect(self.button_new_student_click)

        self.button_normal.clicked.connect(self.button_random_click)

        style_field = "border: 2px solid grey; background: white;  color: black; border-radius: 15px; "
        style_field_button = "border: 2px solid grey; background-color:white; color:black;  border-radius: 10px;"
        #style_field_now = "("background-color: rgba(255, 255, 255, 0);\n"border: none;"

        self.button_instruction.setStyleSheet("background-color: rgba(255, 255, 255, 0);\n" "border: none;")
        #self.button_instruction.setStyleSheet(style_field)


        self.button_enter.setStyleSheet(style_field_button)
        self.button_restart.setStyleSheet(style_field_button)
        self.button_start.setStyleSheet(style_field_button)
        self.button_test_finish.setStyleSheet(style_field_button)
        self.restart_value.setStyleSheet(style_field)
        self.button_OK.setStyleSheet(style_field_button)
        self.button_OK_2.setStyleSheet(style_field_button)
        self.button_new_student.setStyleSheet(style_field_button)
        self.FIO_value.setStyleSheet(style_field)

        self.vzvod_value.setStyleSheet(style_field)

        self.table_0_0.setStyleSheet(style_field)
        self.table_0_0.setAlignment(Qt.AlignCenter)
        self.table_0_1.setStyleSheet(style_field)
        self.table_0_1.setAlignment(Qt.AlignCenter)
        self.table_0_2.setStyleSheet(style_field)
        self.table_0_2.setAlignment(Qt.AlignCenter)
        self.table_1_0.setStyleSheet(style_field)
        self.table_1_0.setAlignment(Qt.AlignCenter)
        self.table_1_1.setStyleSheet(style_field)
        self.table_1_1.setAlignment(Qt.AlignCenter)
        self.table_1_2.setStyleSheet(style_field)
        self.table_1_2.setAlignment(Qt.AlignCenter)
        self.table_2_0.setStyleSheet(style_field)
        self.table_2_0.setAlignment(Qt.AlignCenter)
        self.table_2_1.setStyleSheet(style_field)
        self.table_2_1.setAlignment(Qt.AlignCenter)
        self.table_2_2.setStyleSheet(style_field)
        self.table_2_2.setAlignment(Qt.AlignCenter)
        self.table_2_3.setStyleSheet(style_field)
        self.table_2_3.setAlignment(Qt.AlignCenter)
        self.button_normal.setStyleSheet(style_field)
        self.button_of_restart_variant.setStyleSheet(style_field)
        self.table_0_0.setFont(QFont('Times', 12))
        self.table_0_1.setFont(QFont('Times', 12))
        self.table_0_2.setFont(QFont('Times', 12))
        self.table_1_0.setFont(QFont('Times', 12))
        self.table_1_1.setFont(QFont('Times', 12))
        self.table_1_2.setFont(QFont('Times', 12))
        self.table_2_0.setFont(QFont('Times', 12))
        self.table_2_1.setFont(QFont('Times', 12))
        self.table_2_2.setFont(QFont('Times', 12))
        self.table_2_3.setFont(QFont('Times', 12))
        self.restart_text.setFont(QFont('Times', 12))
        self.restart_text.adjustSize()
        self.russkladka_rus(1)  # смена на рус

    def resizeEvent(self, event):
        print(event)
        self.w = self.size().width()  # "определение ширины"
        self.h = self.size().height()
        print(self.w, self.h)
        center_h = int(self.h / 2)
        center_w = int(self.w / 2)
        self.FIO_value.move(center_w - 80, center_h - 100)
        self.vzvod_value.move(center_w - 80, center_h - 50)
        self.vzvod_text.move(center_w - 200, center_h - 42)
        self.FIO_text.move(center_w - 200, center_h - 92)
        self.button_enter.move(center_w - 80, center_h)
        self.hint_text.move(center_w - 110, center_h - 250)
        self.button_start.move(center_w - 80, center_h + 30)
        self.table_1_1.move(center_w - 115, center_h - 50)


        self.table_0_0.move(center_w - 280, center_h - 115)
        self.table_0_1.move(center_w - 115, center_h - 115)
        self.table_0_2.move(center_w + 50, center_h - 115)
        self.table_1_0.move(center_w - 280, center_h - 50)
        self.table_1_2.move(center_w + 50, center_h - 50)
        self.table_2_0.move(center_w - 280, center_h + 15)
        self.table_2_1.move(center_w - 115, center_h + 15)
        self.table_2_2.move(center_w + 50, center_h + 15)
        self.table_2_3.move(center_w + 240, center_h - 50)
        self.symbol1k.move(center_w + 220, center_h - 40)
        self.mashtab_text.move(center_w + 245, center_h - 70)

        self.button_new_student.move(center_w + 50, center_h + 110)
        self.button_test_finish.move(center_w - 90, center_h + 110)
        self.button_restart.move(center_w - 90, center_h + 110)
        self.restart_value.move(center_w - 115, center_h + 135)
        self.restart_text.move(center_w - 105, center_h + 110)
        self.button_OK.move(center_w - 77, center_h + 200)
        self.button_OK_2.move(center_w - 77, center_h + 200)
        self.button_of_restart_variant.move(center_w - 280, center_h + 110)
        self.button_normal.move(center_w - 115, center_h - 115)  ######
        self.button_instruction.move(self.w-(self.w-20), self.h-50)

        self.w_count = int(0.2 * (1920 - self.w))
        self.h_count = int(0.75 * (1009 - self.h))
        self.pix_gerb = self.pixmap.scaled(500 - self.w_count, 500 - self.h_count)
        self.pix = self.pixmap_wpora.scaled(500 - self.w_count, 500 - self.h_count)
        self.label_wpora.setPixmap(self.pix)
        self.label_wpora.resize(500 - self.w_count, 500 - self.h_count)
        self.label.setPixmap(self.pix_gerb)
        self.label.resize(470 - self.w_count, 470 - self.h_count)
        self.label_wpora.move(300 + (self.w - 800) + self.w_count, 100 + (self.h - 600) + self.h_count)
        # self.pixmap = self.pixmap.scaled(500, 500)
        self.button_instruction.resize(500, 500)
        self.button_instruction.move(300 + (self.w - 800) + self.w_count, 100 + (self.h - 600) + self.h_count)

        self.pix2 = self.pixmap_wpora2.scaled(30, 500 - self.h_count)
        self.label_wpora2.setPixmap(self.pix2)
        self.label_wpora2.resize(30, 500 - self.h_count)
        self.label_wpora2.move(270 + self.w - 800 + self.w_count, 100 + (self.h - 600) + self.h_count)

    def keyPressEvent(self, event):
        if event.key() == 16777220:
            if (self.button_enter.isVisible()):
                self.enter_button()
            elif (self.button_OK.isVisible()):
                self.button_OK_click()
            elif (self.button_OK_2.isVisible()):
                self.button_OK_2_click()
            elif (self.button_start.isVisible()):
                self.button_start_click()

def application() -> None:
    """"Start aplication mainwindow"""
    try:  # это хня позволяет иметь иконку на панеле снизу
        # Включите в блок try/except, если вы также нацелены на Mac/Linux
        from PyQt5.QtWinExtras import QtWin  # !!!
        myappid = 'mycompany.myproduct.subproduct.version'  # !!!
        QtWin.setCurrentProcessExplicitAppUserModelID(myappid)  # !!!
    except ImportError:
        pass

    app = QApplication(sys.argv)
    splash = SplashScreen()
    splash.show()
    splash.progress()
    splash.finish(splash)
    app.setWindowIcon(QtGui.QIcon(resource_path('icons.png')))

    instruction = Instruction()
    instruction.setObjectName("MainWindow")

    instruction.setWindowIcon(QtGui.QIcon(resource_path('icons.png')))

    instruction.setWindowTitle("Инструкция")
    instruction.setMinimumSize(600, 400)
    instruction.setMaximumSize(1000, 800)

    window = Window(instruction)
    window.setObjectName("MainWindow")

    window.setWindowIcon(QtGui.QIcon(resource_path('icons.png')))
    window.setMinimumSize(1000, 730)
    window.setWindowTitle("Номенклатура")

    window.setStyleSheet("#MainWindow{border-image:url(won2.png)}")  # 3e753b

    window.showMaximized()

    sys.exit(app.exec_())


if __name__ == "__main__":
    application()